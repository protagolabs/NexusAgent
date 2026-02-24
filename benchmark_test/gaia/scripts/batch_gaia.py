"""
GAIA Batch Runner
=================
Runs a range of GAIA tasks one-by-one via run_gaia.py with fault tolerance.
If a task crashes, it logs the error and continues to the next one.
Saves JSON + Excel after each question for live monitoring.

Usage:
    python batch_gaia.py --start 2 --end 20 --split validation
    python batch_gaia.py --start 0 --end 50 --split validation --level 1
    python batch_gaia.py --resume results/gaia_validation_batch_2-19_*.json --end 20
"""

import argparse
import json
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
RUN_GAIA = SCRIPT_DIR / "run_gaia.py"
RESULTS_DIR = SCRIPT_DIR.parent / "results"
PYTHON = sys.executable


def load_existing_results(path: str) -> dict:
    with open(path) as f:
        return json.load(f)


def save_json(results: list, split: str, start: int, end: int, correct: int) -> Path:
    """Save consolidated JSON (overwrites same file each time)."""
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    outfile = RESULTS_DIR / f"gaia_{split}_batch_{start}-{end - 1}.json"
    scored = [r for r in results if r.get("correct") is not None]
    output = {
        "metadata": {
            "split": split,
            "range": f"{start}-{end - 1}",
            "timestamp": datetime.now().isoformat(),
            "total_tasks": len(results),
            "scored": len(scored),
            "correct": correct,
        },
        "results": sorted(results, key=lambda r: r.get("task_index", 0)),
    }
    with open(outfile, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    return outfile


def save_excel(results: list, split: str, start: int, end: int, correct: int):
    """Save/overwrite Excel with Summary, Results, and Reasoning Traces sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return  # skip if openpyxl not available

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = RESULTS_DIR / f"gaia_{split}_batch_{start}-{end - 1}.xlsx"

    wb = Workbook()
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    bold = Font(bold=True)

    sorted_results = sorted(results, key=lambda r: r.get("task_index", 0))
    scored = [r for r in sorted_results if r.get("correct") is not None]
    total = len(sorted_results)

    # --- Summary sheet ---
    ws0 = wb.active
    ws0.title = "Summary"
    summary_rows = [
        ("Split", split),
        ("Range", f"{start}-{end - 1}"),
        ("Timestamp", datetime.now().isoformat()),
        ("Total Tasks", total),
        ("Scored", len(scored)),
        ("Correct", correct),
        ("Accuracy", f"{100 * correct / len(scored):.1f}%" if scored else "N/A"),
    ]
    # Per-level stats
    for lvl in [1, 2, 3]:
        lvl_all = [r for r in scored if str(r.get("level", "")) == str(lvl)]
        lvl_ok = sum(1 for r in lvl_all if r.get("correct"))
        if lvl_all:
            summary_rows.append((f"Level {lvl}", f"{lvl_ok}/{len(lvl_all)} ({100 * lvl_ok / len(lvl_all):.1f}%)"))
    for i, (k, v) in enumerate(summary_rows, 1):
        ws0.cell(row=i, column=1, value=k).font = bold
        ws0.cell(row=i, column=2, value=v)
    ws0.column_dimensions["A"].width = 15
    ws0.column_dimensions["B"].width = 40

    # --- Results sheet ---
    ws1 = wb.create_sheet("Results")
    headers = ["Index", "Level", "Correct", "Predicted", "Ground Truth", "Failure", "Duration(s)", "Question (preview)", "Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(sorted_results, 2):
        ok = r.get("correct", False)
        ft = r.get("failure_type", "unknown")
        ws1.cell(row=i, column=1, value=r.get("task_index"))
        ws1.cell(row=i, column=2, value=str(r.get("level", "")))
        ws1.cell(row=i, column=3, value="Yes" if ok else "No")
        ws1.cell(row=i, column=4, value=r.get("predicted_answer", ""))
        ws1.cell(row=i, column=5, value=r.get("ground_truth", ""))
        ws1.cell(row=i, column=6, value=ft)
        ws1.cell(row=i, column=7, value=r.get("duration_seconds"))
        q = r.get("question", "")
        ws1.cell(row=i, column=8, value=q[:150] if q else "")
        ws1.cell(row=i, column=9, value="")
        fill = green if ok else (yellow if ft not in ("success", "") else red)
        for c in range(1, 10):
            ws1.cell(row=i, column=c).fill = fill

    # --- Reasoning Traces sheet ---
    ws2 = wb.create_sheet("Reasoning Traces")
    headers2 = ["Index", "Level", "Correct", "Question", "Response Text", "Thinking Text", "Tool Calls"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h).font = bold
    for i, r in enumerate(sorted_results, 2):
        ws2.cell(row=i, column=1, value=r.get("task_index"))
        ws2.cell(row=i, column=2, value=str(r.get("level", "")))
        ws2.cell(row=i, column=3, value="Yes" if r.get("correct") else "No")
        ws2.cell(row=i, column=4, value=r.get("question", ""))
        ws2.cell(row=i, column=5, value=r.get("response_text", ""))
        ws2.cell(row=i, column=6, value=r.get("thinking_text", ""))
        tools = r.get("tool_calls", [])
        if tools:
            tool_summary = ", ".join(t.get("tool_name", "") for t in tools if isinstance(t, dict))
            ws2.cell(row=i, column=7, value=tool_summary)

    wb.save(xlsx_path)
    return xlsx_path


def run_single(task_id: int, split: str, extra_args: list[str]) -> dict | None:
    """Run a single task via run_gaia.py and return the result dict, or None on crash."""
    cmd = [
        PYTHON, str(RUN_GAIA),
        "--task-id", str(task_id),
        "--split", split,
        *extra_args,
    ]
    try:
        proc = subprocess.run(
            cmd, capture_output=True, text=True, timeout=1800,  # 30 min max per task
        )
    except subprocess.TimeoutExpired:
        return {"task_index": task_id, "failure_type": "batch_timeout",
                "failure_detail": "Task timed out (30 min)", "correct": False}
    except Exception as e:
        return {"task_index": task_id, "failure_type": "batch_crash",
                "failure_detail": str(e), "correct": False}

    # Find the result JSON produced by run_gaia.py
    pattern = f"gaia_{split}_{task_id}-{task_id}_*.json"
    result_files = sorted(RESULTS_DIR.glob(pattern), key=lambda p: p.stat().st_mtime)
    if not result_files:
        return {"task_index": task_id, "failure_type": "no_result_file",
                "failure_detail": proc.stderr[-500:] if proc.stderr else "unknown",
                "correct": False}

    with open(result_files[-1]) as f:
        data = json.load(f)

    if data.get("results"):
        return data["results"][0]
    return None


def print_progress(current: int, total: int, task_id: int, correct: int, done: int,
                   elapsed: float, result: dict | None):
    """Print a trange-style progress line."""
    pct = done / total * 100 if total else 0
    bar_len = 30
    filled = int(bar_len * done / total) if total else 0
    bar = "█" * filled + "░" * (bar_len - filled)

    avg = elapsed / done if done else 0
    eta = avg * (total - done)

    status = ""
    if result:
        icon = "✓" if result.get("correct") else "✗"
        ft = result.get("failure_type", "")
        if ft and ft != "success":
            icon = "!"
        dur = result.get("duration_seconds", 0)
        status = f" | Task {task_id}: {icon} ({dur:.0f}s)"

    print(
        f"\r{pct:5.1f}% |{bar}| {done}/{total} "
        f"[{elapsed:.0f}s<{eta:.0f}s, {correct}✓]{status}",
        end="", flush=True,
    )


def main():
    parser = argparse.ArgumentParser(description="GAIA Batch Runner")
    parser.add_argument("--start", type=int, required=True, help="Start task index (inclusive)")
    parser.add_argument("--end", type=int, required=True, help="End task index (exclusive)")
    parser.add_argument("--split", choices=["validation", "test"], default="validation")
    parser.add_argument("--level", type=int, choices=[1, 2, 3], help="Filter by level")
    parser.add_argument("--resume", type=str, help="Path to previous batch JSON to resume from")
    args = parser.parse_args()

    task_ids = list(range(args.start, args.end))
    total = len(task_ids)

    # Build extra args for run_gaia.py
    extra_args = []
    if args.level:
        extra_args += ["--level", str(args.level)]

    # Load previous results if resuming
    completed = {}
    if args.resume:
        prev = load_existing_results(args.resume)
        for r in prev.get("results", []):
            completed[r["task_index"]] = r
        print(f"Resuming: {len(completed)} tasks already done")

    remaining = [t for t in task_ids if t not in completed]
    results = list(completed.values())
    correct = sum(1 for r in results if r.get("correct"))
    done = len(completed)

    print(f"GAIA Batch: tasks {args.start}-{args.end - 1} ({args.split}), {len(remaining)} to run")

    t0 = time.time()
    print_progress(0, total, -1, correct, done, 0, None)

    for task_id in remaining:
        result = run_single(task_id, args.split, extra_args)
        done += 1
        if result:
            results.append(result)
            if result.get("correct"):
                correct += 1
        elapsed = time.time() - t0
        print_progress(done, total, task_id, correct, done, elapsed, result)

        # Save JSON + Excel after each question
        outfile = save_json(results, args.split, args.start, args.end, correct)
        save_excel(results, args.split, args.start, args.end, correct)

    print()  # newline after progress bar

    # Final summary
    scored = [r for r in results if r.get("correct") is not None]
    print(f"\n{'='*60}")
    print(f"BATCH COMPLETE: {correct}/{len(scored)} correct ({100*correct/len(scored):.1f}%)" if scored else "No scored results")
    durations = [r["duration_seconds"] for r in results if r.get("duration_seconds")]
    if durations:
        print(f"Total time: {sum(durations):.0f}s | Avg: {sum(durations)/len(durations):.0f}s")
    failures = {}
    for r in results:
        ft = r.get("failure_type", "unknown")
        if ft != "success":
            failures[ft] = failures.get(ft, 0) + 1
    if failures:
        print(f"Failures: {failures}")
    print(f"Results: {outfile}")
    xlsx_path = RESULTS_DIR / f"gaia_{args.split}_batch_{args.start}-{args.end - 1}.xlsx"
    print(f"Excel:   {xlsx_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
